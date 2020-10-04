from django.shortcuts import render
from app.models import dataset
from django.http import HttpResponseRedirect, HttpResponse, FileResponse
from django.shortcuts import render, redirect
from telethon import TelegramClient, events, sync
from telethon.tl.types import InputPhoneContact
from telethon import functions, types
from telethon.sync import TelegramClient
import asyncio
import openpyxl
import csv
from django.contrib import messages
from django.shortcuts import get_object_or_404
    
def check(phone_number):
    try:
        phone = '+8801751996786'
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        client = TelegramClient(phone, 1807660, '69168f71fc88455dff4744b0a6adc7bf',loop=loop)
        client.connect()
        contact = InputPhoneContact(client_id = 0, phone = phone_number, first_name="__test__", last_name="__last_test__")
        contacts = client(functions.contacts.ImportContactsRequest([contact]))
        wasonline = contacts.to_dict()['users'][0]['status']['was_online']  
        client.disconnect()
        return wasonline
    except:
        res = "not found"       
        return res

def single(request):
    if request.method == 'GET':
        return render(request,'telegram.html')
    elif request.method == 'POST':
        input = request.POST['q']
            
        phone_number = input
        ress = check(phone_number)
        value=dataset.objects.get(mobile=input)
        value.lastseen = ress
        value.save()
        try:
            data = dataset.objects.filter(mobile__iexact=input)
        except:
            print('NULL')                
        return render(request,'telegram.html',{'show':data})

# Create your views here.
def telegram(request):
    if request.method == 'GET':
        
        return render(request, 'telegram.html')
    elif request.method == 'POST':
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                try:
                    row_data.append(str(cell.value))
                    phonenumber = '{:+}'.format((cell.value))
                    print(phonenumber)
                    ress = check(phonenumber)
                    if ress == 'not found':
                        value=dataset.objects.get(mobile=phonenumber)
                        value.lastseen = ress
                        value.save()
                    else:
                        value=dataset.objects.get(mobile=phonenumber)
                        value.lastseen = ress
                        value.save()
                except:
                    print ("Null")
                    
    messages.success(request, 'Form submission successful')
    return render(request, 'telegram.html')


def export(request):
    response = HttpResponse(content_type='text/csv')

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Last Name', 'NID', 'Last Seen'])

    for member in dataset.objects.all().values_list('firstname', 'lastname', 'nid', 'lastseen'):
        writer.writerow(member)

    response['Content-Disposition'] = 'attachment; filename="members.csv"'

    return response